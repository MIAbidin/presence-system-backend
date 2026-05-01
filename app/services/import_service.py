# app/services/import_service.py
"""
Fase 9 — Import Data Massal
- Import mahasiswa dari Excel
- Import dosen dari Excel
- Generate template Excel
"""
import io
from typing import List, Dict, Any, Optional
from sqlalchemy.orm import Session
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from app.models.user import User, UserRole
from app.services.auth_service import hash_password


# ── TEMPLATE GENERATOR ────────────────────────────────────────

def _base_style(wb: Workbook):
    """Helper style constants."""
    header_font  = Font(bold=True, color="FFFFFF", size=11)
    header_fill  = PatternFill("solid", fgColor="1E3A5F")
    center       = Alignment(horizontal="center", vertical="center")
    left         = Alignment(horizontal="left", vertical="center")
    thin         = Side(style="thin", color="CBD5E1")
    border       = Border(left=thin, right=thin, top=thin, bottom=thin)
    return header_font, header_fill, center, left, border


def generate_template_mahasiswa() -> bytes:
    """Generate template Excel untuk import mahasiswa."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Import Mahasiswa"

    header_font, header_fill, center, left, border = _base_style(wb)

    # Header info
    ws.merge_cells("A1:F1")
    c = ws["A1"]
    c.value     = "TEMPLATE IMPORT DATA MAHASISWA — Presensi SKS"
    c.font      = Font(bold=True, size=13, color="1E3A5F")
    c.alignment = center

    ws.merge_cells("A2:F2")
    c2 = ws["A2"]
    c2.value     = "Isi data mulai baris ke-5. Jangan ubah header kolom. Kolom bertanda * wajib diisi."
    c2.font      = Font(italic=True, size=10, color="64748B")
    c2.alignment = center

    # Header kolom
    headers = [
        ("NIM *",          "Contoh: H071211001", 20),
        ("Nama Lengkap *", "Contoh: Muhammad Rizky Pratama", 30),
        ("Email *",        "Contoh: rizky@student.ac.id", 30),
        ("Password *",     "min. 6 karakter", 20),
        ("Program Studi *","Contoh: Teknik Informatika", 25),
        ("Keterangan",     "Opsional — catatan tambahan", 25),
    ]

    HDR_ROW = 4
    for col, (label, hint, width) in enumerate(headers, start=1):
        # Header
        cell           = ws.cell(row=HDR_ROW, column=col, value=label)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = center
        cell.border    = border

        # Hint row
        hint_cell           = ws.cell(row=HDR_ROW + 1, column=col, value=hint)
        hint_cell.font      = Font(italic=True, size=9, color="94A3B8")
        hint_cell.alignment = left
        hint_cell.border    = border
        hint_cell.fill      = PatternFill("solid", fgColor="F8FAFC")

        ws.column_dimensions[get_column_letter(col)].width = width

    # Contoh data (baris 6)
    examples = [
        ["H071211099", "Contoh Mahasiswa Satu", "contoh1@student.ac.id",
         "Password123!", "Teknik Informatika", ""],
        ["H071411099", "Contoh Mahasiswa Dua", "contoh2@student.ac.id",
         "Password123!", "Sistem Informasi", "Mahasiswa transfer"],
    ]
    for i, row_data in enumerate(examples):
        for col, val in enumerate(row_data, start=1):
            cell           = ws.cell(row=HDR_ROW + 2 + i, column=col, value=val)
            cell.alignment = left
            cell.border    = border
            cell.fill      = PatternFill("solid", fgColor="EFF6FF")

    ws.row_dimensions[HDR_ROW].height = 22
    ws.freeze_panes = f"A{HDR_ROW + 2}"

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def generate_template_dosen() -> bytes:
    """Generate template Excel untuk import dosen."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Import Dosen"

    header_font, header_fill, center, left, border = _base_style(wb)

    ws.merge_cells("A1:F1")
    c = ws["A1"]
    c.value     = "TEMPLATE IMPORT DATA DOSEN — Presensi SKS"
    c.font      = Font(bold=True, size=13, color="1E3A5F")
    c.alignment = center

    ws.merge_cells("A2:F2")
    c2 = ws["A2"]
    c2.value     = "Isi data mulai baris ke-5. Jangan ubah header kolom. Kolom bertanda * wajib diisi."
    c2.font      = Font(italic=True, size=10, color="64748B")
    c2.alignment = center

    headers = [
        ("NIDN *",         "Contoh: 0012038901", 20),
        ("Nama Lengkap *", "Contoh: Dr. Budi Santoso, M.T.", 35),
        ("Email *",        "Contoh: budi@universitashasanuddin.ac.id", 35),
        ("Password *",     "min. 6 karakter", 20),
        ("Program Studi *","Contoh: Teknik Informatika", 25),
        ("Keterangan",     "Opsional — catatan tambahan", 25),
    ]

    HDR_ROW = 4
    for col, (label, hint, width) in enumerate(headers, start=1):
        cell           = ws.cell(row=HDR_ROW, column=col, value=label)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = center
        cell.border    = border

        hint_cell           = ws.cell(row=HDR_ROW + 1, column=col, value=hint)
        hint_cell.font      = Font(italic=True, size=9, color="94A3B8")
        hint_cell.alignment = left
        hint_cell.border    = border
        hint_cell.fill      = PatternFill("solid", fgColor="F8FAFC")

        ws.column_dimensions[get_column_letter(col)].width = width

    examples = [
        ["0012038901", "Dr. Ir. Budi Santoso, M.T.",
         "budi.santoso@universitashasanuddin.ac.id",
         "Password123!", "Teknik Informatika", ""],
        ["0023047802", "Siti Rahayu Ningrum, S.T., M.Sc.",
         "siti.rahayu@universitashasanuddin.ac.id",
         "Password123!", "Sistem Informasi", "Dosen tidak tetap"],
    ]
    for i, row_data in enumerate(examples):
        for col, val in enumerate(row_data, start=1):
            cell           = ws.cell(row=HDR_ROW + 2 + i, column=col, value=val)
            cell.alignment = left
            cell.border    = border
            cell.fill      = PatternFill("solid", fgColor="F0FDF4")

    ws.row_dimensions[HDR_ROW].height = 22
    ws.freeze_panes = f"A{HDR_ROW + 2}"

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


# ── PARSE EXCEL ───────────────────────────────────────────────

def _parse_excel(file_bytes: bytes, role: str) -> List[Dict[str, Any]]:
    """
    Parse file Excel dan return list of dicts.
    Data mulai dari baris 6 (baris 4 = header, 5 = hint, 6+ = data).
    """
    wb   = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws   = wb.active
    rows = []

    DATA_START_ROW = 6  # baris 4=header, 5=hint/contoh, 6+=data asli

    for row_idx, row in enumerate(ws.iter_rows(min_row=DATA_START_ROW, values_only=True), start=DATA_START_ROW):
        # Skip baris kosong
        if not any(row):
            continue
        # Skip baris example jika ada yang masih pakai contoh
        nim_nidn = str(row[0]).strip() if row[0] else ""
        if nim_nidn.lower().startswith("contoh"):
            continue

        rows.append({
            "row_num"      : row_idx,
            "nim_nidn"     : str(row[0]).strip() if row[0] else "",
            "nama_lengkap" : str(row[1]).strip() if row[1] else "",
            "email"        : str(row[2]).strip().lower() if row[2] else "",
            "password"     : str(row[3]).strip() if row[3] else "",
            "program_studi": str(row[4]).strip() if row[4] else "",
            "keterangan"   : str(row[5]).strip() if row[5] else "",
        })

    wb.close()
    return rows


# ── VALIDASI SATU BARIS ───────────────────────────────────────

def _validate_row(row: Dict[str, Any], role: str) -> Optional[str]:
    """Return pesan error jika tidak valid, None jika valid."""
    nim_nidn      = row.get("nim_nidn", "")
    nama_lengkap  = row.get("nama_lengkap", "")
    email         = row.get("email", "")
    password      = row.get("password", "")
    program_studi = row.get("program_studi", "")

    if not nim_nidn:
        return f"{'NIM' if role == 'mahasiswa' else 'NIDN'} wajib diisi"
    if len(nim_nidn) < 5:
        return f"{'NIM' if role == 'mahasiswa' else 'NIDN'} minimal 5 karakter"
    if not nama_lengkap or len(nama_lengkap) < 3:
        return "Nama lengkap minimal 3 karakter"
    if not email or "@" not in email:
        return "Format email tidak valid"
    if not password or len(password) < 6:
        return "Password minimal 6 karakter"
    if not program_studi or len(program_studi) < 3:
        return "Program studi wajib diisi"
    return None


# ── IMPORT MAHASISWA ──────────────────────────────────────────

def import_mahasiswa(db: Session, file_bytes: bytes) -> Dict[str, Any]:
    """
    Parse Excel, validasi, dan insert mahasiswa ke database.
    Return summary: total, berhasil, gagal, errors.
    """
    rows   = _parse_excel(file_bytes, "mahasiswa")
    result = {
        "total"   : len(rows),
        "berhasil": 0,
        "gagal"   : 0,
        "errors"  : [],
        "preview" : [],
    }

    if not rows:
        return {**result, "pesan": "File kosong atau tidak ada data yang valid"}

    # Ambil semua NIM dan email yang sudah ada (bulk query)
    existing_nim    = {u.nim_nidn for u in db.query(User.nim_nidn).all()}
    existing_email  = {u.email for u in db.query(User.email).all()}
    new_nim_set     : set = set()
    new_email_set   : set = set()

    for row in rows:
        row_num = row["row_num"]

        # Validasi format
        err = _validate_row(row, "mahasiswa")
        if err:
            result["gagal"] += 1
            result["errors"].append({"baris": row_num, "nim": row["nim_nidn"], "pesan": err})
            continue

        # Cek duplikat NIM
        if row["nim_nidn"] in existing_nim or row["nim_nidn"] in new_nim_set:
            result["gagal"] += 1
            result["errors"].append({
                "baris": row_num, "nim": row["nim_nidn"],
                "pesan": f"NIM {row['nim_nidn']} sudah terdaftar"
            })
            continue

        # Cek duplikat email
        if row["email"] in existing_email or row["email"] in new_email_set:
            result["gagal"] += 1
            result["errors"].append({
                "baris": row_num, "nim": row["nim_nidn"],
                "pesan": f"Email {row['email']} sudah digunakan"
            })
            continue

        # Insert
        user = User(
            nim_nidn           = row["nim_nidn"],
            nama_lengkap       = row["nama_lengkap"],
            email              = row["email"],
            password_hash      = hash_password(row["password"]),
            role               = UserRole.mahasiswa,
            program_studi      = row["program_studi"],
            is_face_registered = False,
            is_active          = True,
        )
        db.add(user)
        new_nim_set.add(row["nim_nidn"])
        new_email_set.add(row["email"])
        result["berhasil"] += 1

        # Preview 5 baris pertama yang berhasil
        if len(result["preview"]) < 5:
            result["preview"].append({
                "baris"        : row_num,
                "nim_nidn"     : row["nim_nidn"],
                "nama_lengkap" : row["nama_lengkap"],
                "email"        : row["email"],
                "program_studi": row["program_studi"],
            })

    if result["berhasil"] > 0:
        db.commit()

    result["pesan"] = (
        f"{result['berhasil']} mahasiswa berhasil diimport"
        + (f", {result['gagal']} baris gagal" if result["gagal"] > 0 else "")
    )
    return result


# ── IMPORT DOSEN ──────────────────────────────────────────────

def import_dosen(db: Session, file_bytes: bytes) -> Dict[str, Any]:
    """
    Parse Excel, validasi, dan insert dosen ke database.
    """
    rows   = _parse_excel(file_bytes, "dosen")
    result = {
        "total"   : len(rows),
        "berhasil": 0,
        "gagal"   : 0,
        "errors"  : [],
        "preview" : [],
    }

    if not rows:
        return {**result, "pesan": "File kosong atau tidak ada data yang valid"}

    existing_nidn  = {u.nim_nidn for u in db.query(User.nim_nidn).all()}
    existing_email = {u.email for u in db.query(User.email).all()}
    new_nidn_set   : set = set()
    new_email_set  : set = set()

    for row in rows:
        row_num = row["row_num"]

        err = _validate_row(row, "dosen")
        if err:
            result["gagal"] += 1
            result["errors"].append({"baris": row_num, "nidn": row["nim_nidn"], "pesan": err})
            continue

        if row["nim_nidn"] in existing_nidn or row["nim_nidn"] in new_nidn_set:
            result["gagal"] += 1
            result["errors"].append({
                "baris": row_num, "nidn": row["nim_nidn"],
                "pesan": f"NIDN {row['nim_nidn']} sudah terdaftar"
            })
            continue

        if row["email"] in existing_email or row["email"] in new_email_set:
            result["gagal"] += 1
            result["errors"].append({
                "baris": row_num, "nidn": row["nim_nidn"],
                "pesan": f"Email {row['email']} sudah digunakan"
            })
            continue

        user = User(
            nim_nidn           = row["nim_nidn"],
            nama_lengkap       = row["nama_lengkap"],
            email              = row["email"],
            password_hash      = hash_password(row["password"]),
            role               = UserRole.dosen,
            program_studi      = row["program_studi"],
            is_face_registered = False,
            is_active          = True,
        )
        db.add(user)
        new_nidn_set.add(row["nim_nidn"])
        new_email_set.add(row["email"])
        result["berhasil"] += 1

        if len(result["preview"]) < 5:
            result["preview"].append({
                "baris"        : row_num,
                "nim_nidn"     : row["nim_nidn"],
                "nama_lengkap" : row["nama_lengkap"],
                "email"        : row["email"],
                "program_studi": row["program_studi"],
            })

    if result["berhasil"] > 0:
        db.commit()

    result["pesan"] = (
        f"{result['berhasil']} dosen berhasil diimport"
        + (f", {result['gagal']} baris gagal" if result["gagal"] > 0 else "")
    )
    return result


# ── PREVIEW (tanpa insert) ────────────────────────────────────

def preview_excel(file_bytes: bytes, role: str) -> Dict[str, Any]:
    """
    Parse Excel dan return preview 5 baris pertama TANPA insert ke DB.
    Digunakan frontend untuk konfirmasi sebelum import sesungguhnya.
    """
    rows = _parse_excel(file_bytes, role)
    if not rows:
        return {"total": 0, "preview": [], "pesan": "File kosong"}

    preview = []
    for row in rows[:5]:
        err = _validate_row(row, role)
        preview.append({
            **row,
            "valid": err is None,
            "error": err,
        })

    return {
        "total"  : len(rows),
        "preview": preview,
        "pesan"  : f"Ditemukan {len(rows)} baris data",
    }