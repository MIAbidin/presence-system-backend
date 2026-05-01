# app/services/laporan_service.py
"""
Fase 10 — Laporan Global
- GET /admin/laporan              — rekap global dengan filter
- GET /admin/laporan/detail       — detail satu sesi atau semua sesi satu MK
- GET /admin/laporan/export/excel — export ke Excel
- GET /admin/laporan/export/pdf   — export ke PDF
- GET /admin/laporan/mahasiswa/{user_id} — rekap satu mahasiswa
"""

import io
import logging
from datetime import datetime, date
from typing import Optional, List, Dict, Any
from uuid import UUID

from sqlalchemy.orm import Session
from sqlalchemy import func, and_, or_

from app.models.user import User, UserRole
from app.models.matakuliah import Matakuliah
from app.models.sesi import SesiPresensi, SesiStatus, SesiMode
from app.models.presensi import Presensi, PresensiStatus
from app.models.mahasiswa_matakuliah import MahasiswaMatakuliah

logger = logging.getLogger(__name__)

# ── Warna status untuk Excel ──────────────────────────────────
STATUS_COLOR = {
    "hadir"    : "DCFCE7",
    "terlambat": "FEF3C7",
    "absen"    : "FEE2E2",
    "izin"     : "DBEAFE",
    "sakit"    : "F3E8FF",
}

# ── Helper: hitung statistik dari list presensi ────────────────
def _hitung_stat(presensi_list: list) -> Dict:
    total     = len(presensi_list)
    hadir     = sum(1 for p in presensi_list if p.status == PresensiStatus.hadir)
    terlambat = sum(1 for p in presensi_list if p.status == PresensiStatus.terlambat)
    absen     = sum(1 for p in presensi_list if p.status == PresensiStatus.absen)
    izin      = sum(1 for p in presensi_list if p.status == PresensiStatus.izin)
    sakit     = sum(1 for p in presensi_list if p.status == PresensiStatus.sakit)
    efektif   = hadir + terlambat
    persen    = round(efektif / total * 100, 1) if total else 0.0
    return {
        "total"        : total,
        "hadir"        : hadir,
        "terlambat"    : terlambat,
        "absen"        : absen,
        "izin"         : izin,
        "sakit"        : sakit,
        "hadir_efektif": efektif,
        "persentase"   : persen,
    }


def _fmt_time(t) -> Optional[str]:
    if t is None:
        return None
    if hasattr(t, "strftime"):
        return t.strftime("%H:%M")
    return str(t)[:5]


# ════════════════════════════════════════════════════════════
# 10.1 — REKAP GLOBAL
# ════════════════════════════════════════════════════════════

def get_laporan_global(
    db            : Session,
    matakuliah_id : Optional[UUID] = None,
    dosen_id      : Optional[UUID] = None,
    program_studi : Optional[str]  = None,
    periode_mulai : Optional[date] = None,
    periode_selesai: Optional[date]= None,
    mode          : Optional[str]  = None,   # "offline" | "online"
    page          : int = 1,
    limit         : int = 20,
) -> Dict[str, Any]:
    """
    Rekap kehadiran global per matakuliah.
    Setiap baris = satu matakuliah dengan statistik agregat semua sesinya.
    """
    # ── Query sesi dengan filter ──────────────────────────────
    sesi_query = db.query(SesiPresensi).filter(
        SesiPresensi.status == SesiStatus.selesai
    )

    if matakuliah_id:
        sesi_query = sesi_query.filter(SesiPresensi.matakuliah_id == matakuliah_id)

    if dosen_id:
        sesi_query = sesi_query.filter(SesiPresensi.dosen_id == dosen_id)

    if mode:
        try:
            sesi_query = sesi_query.filter(SesiPresensi.mode == SesiMode(mode))
        except ValueError:
            pass

    if periode_mulai:
        sesi_query = sesi_query.filter(
            SesiPresensi.waktu_buka >= datetime.combine(periode_mulai, datetime.min.time())
        )

    if periode_selesai:
        sesi_query = sesi_query.filter(
            SesiPresensi.waktu_buka <= datetime.combine(periode_selesai, datetime.max.time())
        )

    sesi_list = sesi_query.order_by(SesiPresensi.waktu_buka.desc()).all()

    if not sesi_list:
        return {
            "items"      : [],
            "total"      : 0,
            "page"       : page,
            "limit"      : limit,
            "total_pages": 0,
            "ringkasan"  : {"total_sesi": 0, "total_presensi": 0, "rata_rata_kehadiran": 0.0},
        }

    # ── Bulk load presensi untuk semua sesi ───────────────────
    sesi_ids = [s.id for s in sesi_list]
    all_presensi = db.query(Presensi).filter(Presensi.sesi_id.in_(sesi_ids)).all()

    # Group presensi by sesi_id
    presensi_by_sesi: Dict = {}
    for p in all_presensi:
        presensi_by_sesi.setdefault(p.sesi_id, []).append(p)

    # ── Group sesi by matakuliah_id ──────────────────────────
    sesi_by_mk: Dict = {}
    for s in sesi_list:
        sesi_by_mk.setdefault(s.matakuliah_id, []).append(s)

    # ── Bulk load matakuliah & dosen ──────────────────────────
    mk_ids     = list(sesi_by_mk.keys())
    dosen_ids  = list({s.dosen_id for s in sesi_list})

    mk_map = {
        mk.id: mk for mk in
        db.query(Matakuliah).filter(Matakuliah.id.in_(mk_ids)).all()
    }

    dosen_map = {
        u.id: u for u in
        db.query(User).filter(User.id.in_(dosen_ids)).all()
    }

    # ── Filter program_studi via mahasiswa enrolled ───────────
    if program_studi:
        # Filter: hanya tampilkan matakuliah yang ada mahasiswa dari prodi ini
        mhs_mk_rows = db.query(MahasiswaMatakuliah).join(
            User, User.id == MahasiswaMatakuliah.mahasiswa_id
        ).filter(
            func.lower(User.program_studi).like(f"%{program_studi.lower()}%"),
            MahasiswaMatakuliah.matakuliah_id.in_(mk_ids),
        ).all()
        valid_mk_ids = {row.matakuliah_id for row in mhs_mk_rows}
        mk_ids = [mid for mid in mk_ids if mid in valid_mk_ids]
        if not mk_ids:
            return {
                "items": [], "total": 0, "page": page, "limit": limit,
                "total_pages": 0,
                "ringkasan": {"total_sesi": 0, "total_presensi": 0, "rata_rata_kehadiran": 0.0},
            }

    # ── Build result per matakuliah ───────────────────────────
    items_raw = []
    for mk_id in mk_ids:
        mk         = mk_map.get(mk_id)
        sesi_for_mk= sesi_by_mk.get(mk_id, [])
        if not mk or not sesi_for_mk:
            continue

        # Kumpulkan semua presensi di semua sesi MK ini
        all_p = []
        for s in sesi_for_mk:
            all_p.extend(presensi_by_sesi.get(s.id, []))

        stat = _hitung_stat(all_p)

        # Dosen utama (yang paling sering buka sesi)
        dosen_freq: Dict = {}
        for s in sesi_for_mk:
            dosen_freq[s.dosen_id] = dosen_freq.get(s.dosen_id, 0) + 1
        main_dosen_id = max(dosen_freq, key=lambda k: dosen_freq[k])
        dosen_obj     = dosen_map.get(main_dosen_id)

        # Sesi paling terakhir
        sesi_latest = max(sesi_for_mk, key=lambda s: s.waktu_buka or datetime.min)

        items_raw.append({
            "matakuliah_id"     : str(mk_id),
            "kode"              : mk.kode,
            "nama"              : mk.nama,
            "sks"               : mk.sks,
            "hari"              : mk.hari,
            "jam_mulai"         : _fmt_time(mk.jam_mulai),
            "jam_selesai"       : _fmt_time(mk.jam_selesai),
            "ruangan"           : mk.ruangan,
            "dosen_id"          : str(main_dosen_id),
            "nama_dosen"        : dosen_obj.nama_lengkap if dosen_obj else "-",
            "nidn"              : dosen_obj.nim_nidn if dosen_obj else "-",
            "total_pertemuan"   : len(sesi_for_mk),
            "sesi_terakhir"     : sesi_latest.waktu_buka.isoformat() if sesi_latest.waktu_buka else None,
            **stat,
        })

    # Sort: persentase terendah dulu (paling butuh perhatian)
    items_raw.sort(key=lambda x: x["persentase"])

    # ── Pagination manual ─────────────────────────────────────
    total       = len(items_raw)
    total_pages = max(1, (total + limit - 1) // limit)
    start       = (page - 1) * limit
    items_page  = items_raw[start: start + limit]

    # Ringkasan global
    total_sesi      = sum(x["total_pertemuan"] for x in items_raw)
    total_presensi  = sum(x["total"] for x in items_raw)
    avg_kehadiran   = (
        round(sum(x["persentase"] for x in items_raw) / len(items_raw), 1)
        if items_raw else 0.0
    )

    return {
        "items"      : items_page,
        "total"      : total,
        "page"       : page,
        "limit"      : limit,
        "total_pages": total_pages,
        "ringkasan"  : {
            "total_sesi"          : total_sesi,
            "total_presensi"      : total_presensi,
            "rata_rata_kehadiran" : avg_kehadiran,
        },
    }


# ════════════════════════════════════════════════════════════
# 10.2 — DETAIL LAPORAN (per MK atau per sesi)
# ════════════════════════════════════════════════════════════

def get_laporan_detail(
    db           : Session,
    matakuliah_id: Optional[UUID] = None,
    sesi_id      : Optional[UUID] = None,
) -> Dict[str, Any]:
    """
    Detail laporan:
    - Jika sesi_id diisi → detail satu sesi (list mahasiswa + status)
    - Jika matakuliah_id diisi → semua sesi MK tersebut
    """
    if sesi_id:
        return _detail_satu_sesi(db, sesi_id)
    elif matakuliah_id:
        return _detail_per_mk(db, matakuliah_id)
    else:
        return {"error": "Berikan matakuliah_id atau sesi_id"}


def _detail_satu_sesi(db: Session, sesi_id: UUID) -> Dict:
    sesi = db.query(SesiPresensi).filter(SesiPresensi.id == sesi_id).first()
    if not sesi:
        return {"error": "Sesi tidak ditemukan"}

    presensi_list = db.query(Presensi).filter(Presensi.sesi_id == sesi_id).all()
    stat = _hitung_stat(presensi_list)

    detail = []
    for p in presensi_list:
        mhs = p.mahasiswa
        detail.append({
            "presensi_id"   : str(p.id),
            "mahasiswa_id"  : str(p.mahasiswa_id),
            "nim"           : mhs.nim_nidn if mhs else "-",
            "nama_lengkap"  : mhs.nama_lengkap if mhs else "-",
            "program_studi" : mhs.program_studi if mhs else "-",
            "status"        : p.status.value,
            "waktu_presensi": p.waktu_presensi.isoformat() if p.waktu_presensi else None,
            "akurasi_wajah" : p.akurasi_wajah,
            "mode_kelas"    : p.mode_kelas.value,
            "catatan"       : p.catatan,
        })
    detail.sort(key=lambda x: x["nama_lengkap"])

    mk = sesi.matakuliah
    return {
        "tipe"          : "sesi",
        "sesi_id"       : str(sesi_id),
        "matakuliah_id" : str(sesi.matakuliah_id) if sesi.matakuliah_id else None,
        "nama_matakuliah": mk.nama if mk else "-",
        "kode_mk"       : mk.kode if mk else "-",
        "pertemuan_ke"  : sesi.pertemuan_ke,
        "mode"          : sesi.mode.value,
        "waktu_buka"    : sesi.waktu_buka.isoformat() if sesi.waktu_buka else None,
        "waktu_tutup"   : sesi.waktu_tutup.isoformat() if sesi.waktu_tutup else None,
        "statistik"     : stat,
        "detail"        : detail,
    }


def _detail_per_mk(db: Session, matakuliah_id: UUID) -> Dict:
    mk = db.query(Matakuliah).filter(Matakuliah.id == matakuliah_id).first()
    if not mk:
        return {"error": "Matakuliah tidak ditemukan"}

    sesi_list = db.query(SesiPresensi).filter(
        SesiPresensi.matakuliah_id == matakuliah_id
    ).order_by(SesiPresensi.pertemuan_ke).all()

    if not sesi_list:
        return {
            "tipe"          : "matakuliah",
            "matakuliah_id" : str(matakuliah_id),
            "nama"          : mk.nama,
            "kode"          : mk.kode,
            "sesi_list"     : [],
            "statistik_total": _hitung_stat([]),
        }

    sesi_ids     = [s.id for s in sesi_list]
    all_presensi = db.query(Presensi).filter(Presensi.sesi_id.in_(sesi_ids)).all()
    presensi_by_sesi: Dict = {}
    for p in all_presensi:
        presensi_by_sesi.setdefault(p.sesi_id, []).append(p)

    sesi_data = []
    for sesi in sesi_list:
        p_list = presensi_by_sesi.get(sesi.id, [])
        stat   = _hitung_stat(p_list)
        sesi_data.append({
            "sesi_id"     : str(sesi.id),
            "pertemuan_ke": sesi.pertemuan_ke,
            "mode"        : sesi.mode.value,
            "waktu_buka"  : sesi.waktu_buka.isoformat() if sesi.waktu_buka else None,
            "waktu_tutup" : sesi.waktu_tutup.isoformat() if sesi.waktu_tutup else None,
            "status"      : sesi.status.value,
            **stat,
        })

    return {
        "tipe"           : "matakuliah",
        "matakuliah_id"  : str(matakuliah_id),
        "nama"           : mk.nama,
        "kode"           : mk.kode,
        "sesi_list"      : sesi_data,
        "statistik_total": _hitung_stat(all_presensi),
    }


# ════════════════════════════════════════════════════════════
# 10.3 — REKAP PER MAHASISWA
# ════════════════════════════════════════════════════════════

def get_laporan_mahasiswa(db: Session, user_id: UUID) -> Dict[str, Any]:
    """Rekap kehadiran satu mahasiswa di semua matakuliah."""
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        return {"error": "Mahasiswa tidak ditemukan"}

    presensi_all = db.query(Presensi).filter(
        Presensi.mahasiswa_id == user_id
    ).all()

    if not presensi_all:
        return {
            "mahasiswa_id"  : str(user_id),
            "nim"           : user.nim_nidn,
            "nama_lengkap"  : user.nama_lengkap,
            "program_studi" : user.program_studi,
            "total_presensi": 0,
            "statistik"     : _hitung_stat([]),
            "per_matakuliah": [],
        }

    # Bulk load sesi
    sesi_ids  = list({p.sesi_id for p in presensi_all})
    sesi_list = db.query(SesiPresensi).filter(SesiPresensi.id.in_(sesi_ids)).all()
    sesi_map  = {s.id: s for s in sesi_list}

    # Bulk load matakuliah
    mk_ids = list({s.matakuliah_id for s in sesi_list})
    mk_map = {
        mk.id: mk for mk in
        db.query(Matakuliah).filter(Matakuliah.id.in_(mk_ids)).all()
    }

    # Group presensi by matakuliah
    presensi_by_mk: Dict = {}
    for p in presensi_all:
        sesi  = sesi_map.get(p.sesi_id)
        if not sesi:
            continue
        mk_id = sesi.matakuliah_id
        presensi_by_mk.setdefault(mk_id, []).append(p)

    per_matakuliah = []
    for mk_id, p_list in presensi_by_mk.items():
        mk   = mk_map.get(mk_id)
        stat = _hitung_stat(p_list)

        # Riwayat per pertemuan
        riwayat = []
        for p in sorted(p_list, key=lambda x: x.waktu_presensi or datetime.min):
            sesi = sesi_map.get(p.sesi_id)
            riwayat.append({
                "sesi_id"       : str(p.sesi_id),
                "pertemuan_ke"  : sesi.pertemuan_ke if sesi else None,
                "status"        : p.status.value,
                "waktu_presensi": p.waktu_presensi.isoformat() if p.waktu_presensi else None,
                "mode_kelas"    : p.mode_kelas.value,
                "akurasi_wajah" : p.akurasi_wajah,
                "catatan"       : p.catatan,
            })

        per_matakuliah.append({
            "matakuliah_id": str(mk_id),
            "kode"         : mk.kode if mk else "-",
            "nama"         : mk.nama if mk else "-",
            "hari"         : mk.hari if mk else None,
            "jam_mulai"    : _fmt_time(mk.jam_mulai) if mk else None,
            **stat,
            "riwayat"      : riwayat,
        })

    per_matakuliah.sort(key=lambda x: x["persentase"])
    stat_global = _hitung_stat(presensi_all)

    return {
        "mahasiswa_id"  : str(user_id),
        "nim"           : user.nim_nidn,
        "nama_lengkap"  : user.nama_lengkap,
        "program_studi" : user.program_studi,
        "is_face_registered": user.is_face_registered,
        "total_presensi": len(presensi_all),
        "statistik"     : stat_global,
        "per_matakuliah": per_matakuliah,
    }


# ════════════════════════════════════════════════════════════
# 10.4 — EXPORT EXCEL
# ════════════════════════════════════════════════════════════

def export_laporan_excel(
    db            : Session,
    matakuliah_id : Optional[UUID] = None,
    dosen_id      : Optional[UUID] = None,
    program_studi : Optional[str]  = None,
    periode_mulai : Optional[date] = None,
    periode_selesai: Optional[date]= None,
    mode          : Optional[str]  = None,
) -> bytes:
    """Generate file Excel rekap global. Return bytes."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    # Ambil data
    result = get_laporan_global(
        db, matakuliah_id=matakuliah_id, dosen_id=dosen_id,
        program_studi=program_studi, periode_mulai=periode_mulai,
        periode_selesai=periode_selesai, mode=mode,
        page=1, limit=9999,
    )
    items     = result["items"]
    ringkasan = result["ringkasan"]

    wb = Workbook()
    ws = wb.active
    ws.title = "Rekap Laporan"

    # ── Style ────────────────────────────────────────────────
    header_font  = Font(bold=True, color="FFFFFF", size=11)
    header_fill  = PatternFill("solid", fgColor="1E3A5F")
    center       = Alignment(horizontal="center", vertical="center")
    left         = Alignment(horizontal="left",   vertical="center")
    thin         = Side(style="thin", color="CBD5E1")
    border       = Border(left=thin, right=thin, top=thin, bottom=thin)
    title_font   = Font(bold=True, size=14, color="1E3A5F")
    sub_font     = Font(italic=True, size=10, color="64748B")

    # ── Header info ──────────────────────────────────────────
    ws.merge_cells("A1:J1")
    c = ws["A1"]
    c.value     = "LAPORAN REKAP KEHADIRAN — SISTEM PRESENSI"
    c.font      = title_font
    c.alignment = center

    ws.merge_cells("A2:J2")
    parts = []
    if periode_mulai:
        parts.append(f"Periode: {periode_mulai.strftime('%d %b %Y')}")
    if periode_selesai:
        parts.append(f"s/d {periode_selesai.strftime('%d %b %Y')}")
    if mode:
        parts.append(f"Mode: {mode.upper()}")
    ws["A2"].value     = "  ·  ".join(parts) if parts else "Semua Periode"
    ws["A2"].font      = sub_font
    ws["A2"].alignment = center

    # Ringkasan
    ws["A4"].value = "Ringkasan:"
    ws["A4"].font  = Font(bold=True, size=10)
    ws["B4"].value = f"Total Sesi: {ringkasan['total_sesi']}"
    ws["C4"].value = f"Total Presensi: {ringkasan['total_presensi']}"
    ws["D4"].value = f"Rata-rata Kehadiran: {ringkasan['rata_rata_kehadiran']}%"

    # ── Header tabel ─────────────────────────────────────────
    HDR_ROW = 6
    headers = [
        "No", "Kode MK", "Nama Matakuliah", "SKS", "Dosen",
        "Jml Pertemuan", "Hadir", "Terlambat", "Absen", "Izin", "Sakit", "% Kehadiran"
    ]
    for col, h in enumerate(headers, start=1):
        cell           = ws.cell(row=HDR_ROW, column=col, value=h)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = center
        cell.border    = border

    # ── Data ─────────────────────────────────────────────────
    PERSEN_FILL = {
        "good"   : PatternFill("solid", fgColor="DCFCE7"),
        "warning": PatternFill("solid", fgColor="FEF3C7"),
        "danger" : PatternFill("solid", fgColor="FEE2E2"),
    }

    for i, item in enumerate(items, start=1):
        row = HDR_ROW + i
        p   = item["persentase"]
        fill = PERSEN_FILL["good"] if p >= 75 else (PERSEN_FILL["warning"] if p >= 50 else PERSEN_FILL["danger"])

        values = [
            i,
            item["kode"],
            item["nama"],
            item["sks"],
            item["nama_dosen"],
            item["total_pertemuan"],
            item["hadir"],
            item["terlambat"],
            item["absen"],
            item["izin"],
            item["sakit"],
            f"{p}%",
        ]
        aligns = [center, center, left, center, left, center, center, center, center, center, center, center]

        for col, (val, aln) in enumerate(zip(values, aligns), start=1):
            cell           = ws.cell(row=row, column=col, value=val)
            cell.border    = border
            cell.alignment = aln
            # Warnai kolom persentase
            if col == 12:
                cell.fill = fill
                cell.font = Font(bold=True, size=10)

    # ── Lebar kolom ──────────────────────────────────────────
    widths = [5, 10, 32, 5, 28, 14, 8, 12, 8, 8, 8, 14]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.row_dimensions[HDR_ROW].height = 22
    ws.freeze_panes = f"A{HDR_ROW + 1}"

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


# ════════════════════════════════════════════════════════════
# 10.5 — EXPORT PDF
# ════════════════════════════════════════════════════════════

def export_laporan_pdf(
    db            : Session,
    matakuliah_id : Optional[UUID] = None,
    dosen_id      : Optional[UUID] = None,
    program_studi : Optional[str]  = None,
    periode_mulai : Optional[date] = None,
    periode_selesai: Optional[date]= None,
    mode          : Optional[str]  = None,
) -> bytes:
    """
    Generate laporan PDF menggunakan reportlab.
    Return: bytes (PDF file content).
    """
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import (
        SimpleDocTemplate, Table, TableStyle, Paragraph,
        Spacer, HRFlowable,
    )
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

    # Ambil data
    result = get_laporan_global(
        db, matakuliah_id=matakuliah_id, dosen_id=dosen_id,
        program_studi=program_studi, periode_mulai=periode_mulai,
        periode_selesai=periode_selesai, mode=mode,
        page=1, limit=9999,
    )
    items     = result["items"]
    ringkasan = result["ringkasan"]

    buffer   = io.BytesIO()
    doc      = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        leftMargin=1.5*cm, rightMargin=1.5*cm,
        topMargin=1.5*cm,  bottomMargin=1.5*cm,
    )
    styles   = getSampleStyleSheet()
    story    = []

    # ── Warna ────────────────────────────────────────────────
    NAVY     = colors.HexColor("#1E3A5F")
    LIGHT_BG = colors.HexColor("#F8FAFC")
    GREEN    = colors.HexColor("#22c55e")
    AMBER    = colors.HexColor("#f59e0b")
    RED      = colors.HexColor("#ef4444")

    # ── Judul ────────────────────────────────────────────────
    title_style = ParagraphStyle(
        "title", parent=styles["Title"],
        fontSize=16, textColor=NAVY, spaceAfter=4,
        alignment=TA_CENTER, fontName="Helvetica-Bold"
    )
    sub_style = ParagraphStyle(
        "sub", parent=styles["Normal"],
        fontSize=9, textColor=colors.HexColor("#64748B"),
        alignment=TA_CENTER, spaceAfter=12,
    )
    story.append(Paragraph("LAPORAN REKAP KEHADIRAN", title_style))

    parts = []
    if periode_mulai:
        parts.append(f"Periode: {periode_mulai.strftime('%d %b %Y')}")
    if periode_selesai:
        parts.append(f"s/d {periode_selesai.strftime('%d %b %Y')}")
    if mode:
        parts.append(f"Mode: {mode.upper()}")
    story.append(Paragraph(
        "  ·  ".join(parts) if parts else "Semua Periode",
        sub_style
    ))
    story.append(Paragraph(
        f"Dicetak: {datetime.now().strftime('%d %B %Y %H:%M')} WIB",
        sub_style
    ))
    story.append(HRFlowable(width="100%", color=NAVY, thickness=1))
    story.append(Spacer(1, 0.3*cm))

    # ── Ringkasan strip ───────────────────────────────────────
    ring_data = [[
        f"Total Matakuliah: {len(items)}",
        f"Total Sesi: {ringkasan['total_sesi']}",
        f"Total Presensi: {ringkasan['total_presensi']}",
        f"Rata-rata Kehadiran: {ringkasan['rata_rata_kehadiran']}%",
    ]]
    ring_table = Table(ring_data, colWidths=[6*cm, 5*cm, 5*cm, 6*cm])
    ring_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), LIGHT_BG),
        ("TEXTCOLOR",  (0, 0), (-1, -1), NAVY),
        ("FONTNAME",   (0, 0), (-1, -1), "Helvetica-Bold"),
        ("FONTSIZE",   (0, 0), (-1, -1), 9),
        ("ALIGN",      (0, 0), (-1, -1), "CENTER"),
        ("ROWBACKGROUNDS", (0, 0), (-1, -1), [LIGHT_BG]),
        ("TOPPADDING",    (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
    ]))
    story.append(ring_table)
    story.append(Spacer(1, 0.4*cm))

    if not items:
        story.append(Paragraph("Tidak ada data untuk ditampilkan.", styles["Normal"]))
    else:
        # ── Tabel data ────────────────────────────────────────
        table_data = [[
            "No", "Kode", "Nama Matakuliah", "Dosen",
            "Pertemuan", "Hadir", "Terlambat", "Absen", "Izin", "Sakit", "% Hadir"
        ]]

        for i, item in enumerate(items, start=1):
            p = item["persentase"]
            table_data.append([
                str(i),
                item["kode"],
                item["nama"],
                item["nama_dosen"],
                str(item["total_pertemuan"]),
                str(item["hadir"]),
                str(item["terlambat"]),
                str(item["absen"]),
                str(item["izin"]),
                str(item["sakit"]),
                f"{p}%",
            ])

        col_widths = [1*cm, 2.2*cm, 6.5*cm, 6*cm, 2.2*cm, 1.8*cm, 2.2*cm, 1.8*cm, 1.8*cm, 1.8*cm, 2.2*cm]
        tbl = Table(table_data, colWidths=col_widths, repeatRows=1)

        # Base style
        style = [
            ("BACKGROUND",    (0, 0), (-1, 0), NAVY),
            ("TEXTCOLOR",     (0, 0), (-1, 0), colors.white),
            ("FONTNAME",      (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE",      (0, 0), (-1, 0), 8),
            ("ALIGN",         (0, 0), (-1, -1), "CENTER"),
            ("ALIGN",         (2, 1), (3, -1), "LEFT"),
            ("FONTSIZE",      (0, 1), (-1, -1), 7.5),
            ("FONTNAME",      (0, 1), (-1, -1), "Helvetica"),
            ("ROWBACKGROUNDS",(0, 1), (-1, -1), [colors.white, LIGHT_BG]),
            ("GRID",          (0, 0), (-1, -1), 0.5, colors.HexColor("#CBD5E1")),
            ("TOPPADDING",    (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]

        # Warnai baris berdasarkan persentase
        for i, item in enumerate(items, start=1):
            p = item["persentase"]
            if p >= 75:
                bg = colors.HexColor("#DCFCE7")
            elif p >= 50:
                bg = colors.HexColor("#FEF3C7")
            else:
                bg = colors.HexColor("#FEE2E2")
            style.append(("BACKGROUND", (10, i), (10, i), bg))
            style.append(("FONTNAME",   (10, i), (10, i), "Helvetica-Bold"))

        tbl.setStyle(TableStyle(style))
        story.append(tbl)

    # ── Footer ────────────────────────────────────────────────
    story.append(Spacer(1, 0.5*cm))
    story.append(HRFlowable(width="100%", color=colors.HexColor("#CBD5E1"), thickness=0.5))
    story.append(Paragraph(
        "Sistem Presensi Face Recognition — Dokumen ini digenerate otomatis",
        ParagraphStyle("footer", parent=styles["Normal"],
                       fontSize=7, textColor=colors.HexColor("#94A3B8"),
                       alignment=TA_CENTER)
    ))

    doc.build(story)
    return buffer.getvalue()