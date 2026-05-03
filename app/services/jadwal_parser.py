# app/services/jadwal_parser.py
"""
Fase C — Parser jadwal kuliah dari PDF dan Excel.

Mendukung dua format input:
1. PDF   — menggunakan pdfplumber untuk ekstrak tabel
2. Excel (.xlsx/.xls) — menggunakan openpyxl / xlsx_parser

Output: list[dict] dengan struktur:
  {
    hari        : str | None,  # 'Senin', 'Selasa', dst
    kode_ruangan: str | None,  # 'J.Int.1', 'LABRPL', dst
    slot_str    : str | None,  # '1-3', '7-8'
    slot_mulai  : int | None,
    slot_selesai: int | None,
    kode_mk     : str | None,
    nama_mk     : str | None,
    dosen       : str | None,  # nama dosen (raw, belum matched ke DB)
    kode_kelas  : str | None,  # 'A', 'B', 'C', 'X'
    kode_akses  : str | None,  # URL atau kode WA
    _raw_row    : list,        # baris asli untuk debugging
  }

Catatan format PDF kampus:
- Header hari ("Senin", "Selasa", dst) bisa ada di kolom mana saja
- Baris bisa ter-merge (pdfplumber akan return None untuk sel merged)
- Kode MK bisa ada spasi atau newline: "TIF3221308" atau "TIF 3221308"
- Slot bisa format "1-3", "1 - 3", atau "Slot 1-3"
- Satu baris bisa punya beberapa kelas (A, B) dipisah newline
"""

import io
import re
import logging
from typing import Optional

logger = logging.getLogger(__name__)

HARI_LIST = ["Senin", "Selasa", "Rabu", "Kamis", "Jumat", "Sabtu", "Minggu"]

# Pattern untuk kode MK kampus (bisa disesuaikan)
# Contoh: TIF3221308, SI2234567, IK1234567, B10A123
KODE_MK_PATTERN = re.compile(
    r'\b([A-Z]{2,4}\d{5,10})\b',
    re.IGNORECASE
)

# Pattern slot: "1-3", "1 - 3", "Slot 1-3", "1–3"
SLOT_PATTERN = re.compile(
    r'(?:slot\s*)?(\d{1,2})\s*[-–]\s*(\d{1,2})',
    re.IGNORECASE
)

# Pattern kelas: "Kelas A", "Kls B", atau huruf kapital tunggal di kolom kelas
KELAS_PATTERN = re.compile(r'\b(?:kelas\s*|kls\s*)?([A-Z])\b')

# Separator untuk multi-value dalam satu sel
CELL_SEP = re.compile(r'[\n\r;,/]')


# ─── HELPER ───────────────────────────────────────────────────

def _clean(val) -> Optional[str]:
    """Clean cell value dari PDF — handle None, strip whitespace."""
    if val is None:
        return None
    s = str(val).strip()
    # Hapus karakter kontrol tapi pertahankan newline untuk deteksi multi-value
    s = re.sub(r'[\r\t\x00-\x08\x0b\x0c\x0e-\x1f\x7f]', ' ', s)
    s = re.sub(r' +', ' ', s).strip()
    return s if s else None


def _parse_slot(text: str) -> tuple[Optional[int], Optional[int]]:
    """
    Parse string slot ke (slot_mulai, slot_selesai).
    Contoh:
    - "1-3"    → (1, 3)
    - "7 - 8"  → (7, 8)
    - "Slot 4" → (4, 4)   # satu slot
    - "10-12"  → (10, 12)
    """
    if not text:
        return None, None
    text = text.strip()
    m = SLOT_PATTERN.search(text)
    if m:
        s, e = int(m.group(1)), int(m.group(2))
        if 1 <= s <= 12 and 1 <= e <= 12 and s <= e:
            return s, e
    # Single slot: angka tunggal
    single = re.search(r'\b(\d{1,2})\b', text)
    if single:
        n = int(single.group(1))
        if 1 <= n <= 12:
            return n, n
    return None, None


def _detect_hari(text: str) -> Optional[str]:
    """Deteksi nama hari dari teks."""
    if not text:
        return None
    t = text.strip().capitalize()
    for hari in HARI_LIST:
        if hari.lower() in text.lower():
            return hari
    return None


def _extract_kode_mk(text: str) -> Optional[str]:
    """Ekstrak kode MK dari teks."""
    if not text:
        return None
    m = KODE_MK_PATTERN.search(text)
    if m:
        return m.group(1).upper().replace(' ', '')
    # Fallback: ambil token alfanumerik yang mirip kode MK
    tokens = re.findall(r'[A-Z]{2,4}\d{4,}', text.upper())
    return tokens[0] if tokens else None


def _extract_kode_kelas(text: str) -> Optional[str]:
    """Ekstrak kode kelas (A, B, C, X) dari teks."""
    if not text:
        return None
    # Cari pola "Kelas A" atau "A" atau "Kls B"
    m = KELAS_PATTERN.search(text)
    if m:
        return m.group(1).upper()
    # Fallback: ambil huruf kapital tunggal
    singles = re.findall(r'\b([A-Z])\b', text.upper())
    if singles:
        return singles[0]
    return None


def _build_row(
    hari: Optional[str],
    kode_ruangan: Optional[str],
    slot_str: Optional[str],
    kode_mk: Optional[str],
    nama_mk: Optional[str],
    dosen: Optional[str],
    kode_kelas: Optional[str],
    kode_akses: Optional[str],
    raw_row: list,
) -> Optional[dict]:
    """Build result dict — return None jika tidak ada data minimal."""
    if not kode_mk and not nama_mk:
        return None

    slot_mulai, slot_selesai = _parse_slot(slot_str or "")

    return {
        "hari"        : hari,
        "kode_ruangan": kode_ruangan,
        "slot_str"    : slot_str,
        "slot_mulai"  : slot_mulai,
        "slot_selesai": slot_selesai,
        "kode_mk"     : kode_mk,
        "nama_mk"     : nama_mk,
        "dosen"       : dosen,
        "kode_kelas"  : kode_kelas,
        "kode_akses"  : kode_akses,
        "_raw_row"    : raw_row,
    }


# ─── PDF PARSER ───────────────────────────────────────────────

def parse_pdf_jadwal(file_bytes: bytes) -> list[dict]:
    """
    Parse PDF jadwal kampus menggunakan pdfplumber.

    Strategi:
    1. Iterasi setiap halaman, ekstrak semua tabel
    2. Deteksi baris header hari (Senin, Selasa, dst)
    3. Untuk setiap baris data, coba parse kolom-kolom standar
    4. Handle sel merged (None) dengan carry-over dari baris sebelumnya

    Return list of parsed dicts, sudah difilter yang None.
    """
    try:
        import pdfplumber
    except ImportError:
        raise ImportError(
            "pdfplumber belum terinstall. Jalankan: pip install pdfplumber"
        )

    results = []
    current_hari    = None
    prev_ruangan    = None
    prev_slot       = None

    try:
        with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
            for page_num, page in enumerate(pdf.pages, start=1):
                logger.debug(f"Parsing page {page_num}")

                # Coba ekstrak tabel dari halaman
                tables = page.extract_tables({
                    "vertical_strategy"  : "lines",
                    "horizontal_strategy": "lines",
                    "snap_tolerance"     : 5,
                })

                if not tables:
                    # Fallback: ekstrak text biasa
                    text = page.extract_text() or ""
                    rows = _parse_text_fallback(text, current_hari)
                    results.extend(rows)
                    continue

                for table in tables:
                    if not table:
                        continue

                    for row in table:
                        if not row:
                            continue

                        cleaned = [_clean(c) for c in row]

                        # Deteksi baris header hari
                        full_text = " ".join(c for c in cleaned if c)
                        hari_detected = _detect_hari(full_text)
                        if hari_detected and _is_header_row(cleaned):
                            current_hari = hari_detected
                            logger.debug(f"Hari terdeteksi: {current_hari}")
                            continue

                        # Parse baris data
                        parsed = _parse_row_pdf(
                            cleaned,
                            current_hari,
                            prev_ruangan,
                            prev_slot,
                        )

                        if parsed:
                            results.extend(parsed)
                            # Update carry-over
                            if parsed[-1].get("kode_ruangan"):
                                prev_ruangan = parsed[-1]["kode_ruangan"]
                            if parsed[-1].get("slot_str"):
                                prev_slot = parsed[-1]["slot_str"]

    except Exception as e:
        logger.error(f"Error parsing PDF: {e}", exc_info=True)
        raise ValueError(f"Gagal membaca file PDF: {str(e)}")

    return [r for r in results if r is not None]


def _is_header_row(cells: list) -> bool:
    """Cek apakah baris ini adalah header (bukan data)."""
    text = " ".join(c for c in cells if c).lower()
    header_keywords = ["hari", "ruang", "jam", "slot", "matakuliah", "dosen", "kelas"]
    matches = sum(1 for kw in header_keywords if kw in text)
    return matches >= 2


def _parse_row_pdf(
    cells: list,
    current_hari: Optional[str],
    prev_ruangan: Optional[str],
    prev_slot: Optional[str],
) -> list[dict]:
    """
    Parse satu baris tabel PDF.

    Kolom yang diharapkan (urutan bisa berbeda antar kampus):
    [Ruang, Slot/Jam, Kode MK, Nama MK, Dosen, Kelas, Kode Akses]

    Return list karena satu baris bisa punya multi kelas.
    """
    if not any(cells):
        return []

    full_text = " ".join(c for c in cells if c)

    # Cek apakah ada data bermakna
    kode_mk = _extract_kode_mk(full_text)
    if not kode_mk:
        # Cari nama MK dari sel yang cukup panjang
        nama_mk_candidates = [c for c in cells if c and len(c) > 8 and not re.search(r'\d{5,}', c)]
        if not nama_mk_candidates:
            return []

    results = []

    # Deteksi kolom berdasarkan posisi dan konten
    kode_ruangan = _find_ruangan(cells) or prev_ruangan
    slot_str     = _find_slot(cells) or prev_slot
    nama_mk      = _find_nama_mk(cells)
    kode_mk      = _find_kode_mk(cells) or kode_mk
    dosen        = _find_dosen(cells)
    kode_akses   = _find_kode_akses(cells)

    # Cari kelas — bisa multi-kelas dalam satu baris
    kelas_list = _find_kelas_list(cells)
    if not kelas_list:
        kelas_list = [None]

    for kode_kelas in kelas_list:
        row = _build_row(
            hari        = current_hari,
            kode_ruangan= kode_ruangan,
            slot_str    = slot_str,
            kode_mk     = kode_mk,
            nama_mk     = nama_mk,
            dosen       = dosen,
            kode_kelas  = kode_kelas,
            kode_akses  = kode_akses,
            raw_row     = cells,
        )
        if row:
            results.append(row)

    return results


def _find_ruangan(cells: list) -> Optional[str]:
    """Cari kode ruangan dari sel-sel."""
    for cell in cells:
        if not cell:
            continue
        # Pola ruangan: J.Int.1, LABRPL, J0403, LSITIF, SW706
        m = re.search(r'\b([A-Z]{1,4}[\.\-]?[A-Z0-9]{2,8})\b', cell.upper())
        if m and not re.search(r'\d{6,}', m.group(1)):
            return m.group(1)
    return None


def _find_slot(cells: list) -> Optional[str]:
    """Cari string slot dari sel-sel."""
    for cell in cells:
        if not cell:
            continue
        m = SLOT_PATTERN.search(cell)
        if m:
            return f"{m.group(1)}-{m.group(2)}"
    return None


def _find_kode_mk(cells: list) -> Optional[str]:
    """Cari kode MK dari sel-sel."""
    for cell in cells:
        if not cell:
            continue
        m = KODE_MK_PATTERN.search(cell)
        if m:
            return m.group(1).upper().replace(' ', '')
    return None


def _find_nama_mk(cells: list) -> Optional[str]:
    """Cari nama MK (string panjang yang bukan angka/kode)."""
    candidates = []
    for cell in cells:
        if not cell:
            continue
        # Nama MK biasanya: >10 karakter, banyak huruf, sedikit angka
        if len(cell) > 8 and re.search(r'[a-zA-Z\s]{6,}', cell):
            # Bukan slot, bukan kode MK
            if not SLOT_PATTERN.search(cell) and len(re.findall(r'\d', cell)) < len(cell) * 0.5:
                candidates.append(cell)
    return candidates[0] if candidates else None


def _find_dosen(cells: list) -> Optional[str]:
    """Cari nama dosen — biasanya string dengan gelar atau huruf kapital."""
    for cell in cells:
        if not cell:
            continue
        # Cek pola nama dosen: ada kata dengan huruf kapital, mungkin ada Dr./Prof./M.T.
        if re.search(r'\b(Dr|Prof|Ir|M\.T|M\.Sc|M\.Kom|S\.T|S\.Kom)\b', cell, re.IGNORECASE):
            return cell.strip()
        # Nama orang: minimal 2 kata, tidak ada angka banyak
        words = [w for w in cell.split() if len(w) > 1]
        if len(words) >= 2 and len(re.findall(r'\d', cell)) < 2:
            if cell[0].isupper() and len(cell) > 5:
                return cell.strip()
    return None


def _find_kode_akses(cells: list) -> Optional[str]:
    """Cari kode akses (URL atau kode WA) dari sel-sel."""
    for cell in cells:
        if not cell:
            continue
        if re.search(r'https?://', cell, re.IGNORECASE):
            return cell.strip()
        if re.search(r'classroom\.google\.com|wa\.me|bit\.ly', cell, re.IGNORECASE):
            return cell.strip()
    return None


def _find_kelas_list(cells: list) -> list[Optional[str]]:
    """
    Cari daftar kelas dari sel-sel.
    Bisa return ['A'] atau ['A', 'B'] jika multi-kelas dalam satu baris.
    """
    all_kelas = []
    for cell in cells:
        if not cell:
            continue
        # Cari semua huruf kelas dalam sel (A, B, C, X, dst)
        parts = CELL_SEP.split(cell)
        for part in parts:
            m = KELAS_PATTERN.search(part.strip())
            if m and m.group(1) not in all_kelas:
                all_kelas.append(m.group(1).upper())
    return all_kelas if all_kelas else []


def _parse_text_fallback(text: str, current_hari: Optional[str]) -> list[dict]:
    """
    Fallback parser jika pdfplumber tidak bisa ekstrak tabel.
    Parse baris teks biasa.
    """
    results = []
    lines = text.split('\n')
    hari = current_hari

    for line in lines:
        line = line.strip()
        if not line:
            continue

        # Deteksi hari
        h = _detect_hari(line)
        if h and len(line) < 20:
            hari = h
            continue

        kode_mk = _extract_kode_mk(line)
        if kode_mk:
            slot_str = None
            m = SLOT_PATTERN.search(line)
            if m:
                slot_str = f"{m.group(1)}-{m.group(2)}"

            row = _build_row(
                hari        = hari,
                kode_ruangan= _find_ruangan([line]),
                slot_str    = slot_str,
                kode_mk     = kode_mk,
                nama_mk     = None,
                dosen       = None,
                kode_kelas  = _extract_kode_kelas(line),
                kode_akses  = None,
                raw_row     = [line],
            )
            if row:
                results.append(row)

    return results


# ─── EXCEL PARSER ─────────────────────────────────────────────

def parse_excel_jadwal(file_bytes: bytes) -> list[dict]:
    """
    Parse Excel jadwal kampus.

    Mendukung format:
    1. Template standar (dari /admin/import/template/jadwal)
    2. Excel jadwal kampus asli (format bisa berbeda)

    Return list of parsed dicts.
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise ImportError("openpyxl belum terinstall")

    results = []
    current_hari = None

    try:
        wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        ws = wb.active

        for row in ws.iter_rows(values_only=True):
            if not row or not any(row):
                continue

            cleaned = [_clean(v) for v in row]
            full_text = " ".join(c for c in cleaned if c)

            # Deteksi header hari
            hari_detected = _detect_hari(full_text)
            if hari_detected and _is_header_row(cleaned):
                current_hari = hari_detected
                continue

            # Cek apakah baris adalah header tabel
            if _is_header_row(cleaned):
                continue

            parsed = _parse_row_excel(cleaned, current_hari)
            results.extend(parsed)

        wb.close()

    except Exception as e:
        logger.error(f"Error parsing Excel: {e}", exc_info=True)
        raise ValueError(f"Gagal membaca file Excel: {str(e)}")

    return [r for r in results if r is not None]


def _parse_row_excel(cells: list, current_hari: Optional[str]) -> list[dict]:
    """
    Parse satu baris Excel jadwal.

    Untuk template standar, kolom: Hari, Ruangan, Slot, Kode MK, Nama MK, Dosen, Kelas, Kode Akses
    """
    if not any(cells):
        return []

    # Coba format template standar (8 kolom)
    if len(cells) >= 6:
        hari_col    = cells[0] if cells[0] else current_hari
        ruangan_col = cells[1] if len(cells) > 1 else None
        slot_col    = cells[2] if len(cells) > 2 else None
        kode_mk_col = cells[3] if len(cells) > 3 else None
        nama_mk_col = cells[4] if len(cells) > 4 else None
        dosen_col   = cells[5] if len(cells) > 5 else None
        kelas_col   = cells[6] if len(cells) > 6 else None
        akses_col   = cells[7] if len(cells) > 7 else None

        hari = _detect_hari(hari_col or "") or current_hari

        # Jika hari di kolom pertama, update current hari
        if hari and hari_col:
            current_hari = hari

        kode_mk = kode_mk_col or _extract_kode_mk(" ".join(c for c in cells if c))
        if not kode_mk:
            return []

        slot_str = slot_col
        kelas_list = []
        if kelas_col:
            for part in CELL_SEP.split(kelas_col):
                m = KELAS_PATTERN.search(part.strip())
                if m:
                    kelas_list.append(m.group(1).upper())

        if not kelas_list:
            kelas_list = [None]

        rows = []
        for kode_kelas in kelas_list:
            row = _build_row(
                hari        = hari,
                kode_ruangan= ruangan_col,
                slot_str    = slot_str,
                kode_mk     = kode_mk,
                nama_mk     = nama_mk_col,
                dosen       = dosen_col,
                kode_kelas  = kode_kelas,
                kode_akses  = akses_col,
                raw_row     = cells,
            )
            if row:
                rows.append(row)
        return rows

    # Fallback: parse bebas dari semua sel
    full_text = " ".join(c for c in cells if c)
    kode_mk = _extract_kode_mk(full_text)
    if not kode_mk:
        return []

    row = _build_row(
        hari        = current_hari,
        kode_ruangan= _find_ruangan(cells),
        slot_str    = _find_slot(cells),
        kode_mk     = kode_mk,
        nama_mk     = _find_nama_mk(cells),
        dosen       = _find_dosen(cells),
        kode_kelas  = _extract_kode_kelas(full_text),
        kode_akses  = _find_kode_akses(cells),
        raw_row     = cells,
    )
    return [row] if row else []


# ─── TEMPLATE GENERATOR ───────────────────────────────────────

def generate_template_jadwal() -> bytes:
    """
    Generate template Excel untuk import jadwal.
    Format: 8 kolom standar yang bisa diisi manual.
    """
    import io as _io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Template Jadwal"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="1E3A5F")
    center      = Alignment(horizontal="center", vertical="center")
    left        = Alignment(horizontal="left",   vertical="center")
    thin        = Side(style="thin", color="CBD5E1")
    border      = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Judul
    ws.merge_cells("A1:H1")
    c = ws["A1"]
    c.value     = "TEMPLATE IMPORT JADWAL KULIAH — Presensi SKS"
    c.font      = Font(bold=True, size=13, color="1E3A5F")
    c.alignment = center

    ws.merge_cells("A2:H2")
    c2 = ws["A2"]
    c2.value     = "Isi data mulai baris ke-5. Kolom Hari bisa diisi sekali per blok, sisanya biarkan kosong (akan carry-over). Kolom bertanda * wajib diisi."
    c2.font      = Font(italic=True, size=9, color="64748B")
    c2.alignment = center

    headers = [
        ("Hari",       "Senin/Selasa/Rabu/Kamis/Jumat/Sabtu",  15),
        ("Ruangan *",  "J.Int.1 / LABRPL / J0403",              18),
        ("Slot *",     "Format: '1-3' atau '7-8'",              12),
        ("Kode MK *",  "TIF3221308, SI2234567",                 16),
        ("Nama MK",    "Pemrograman Mobile, Basis Data",         30),
        ("Dosen",      "Nama dosen (sesuai di sistem)",          30),
        ("Kelas *",    "A / B / C / X (bisa lebih dari 1, pisah koma)", 10),
        ("Kode Akses", "URL Google Classroom atau kode WA",     35),
    ]

    HDR_ROW = 4
    for col, (label, hint, width) in enumerate(headers, start=1):
        cell           = ws.cell(row=HDR_ROW, column=col, value=label)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = center
        cell.border    = border

        hint_cell           = ws.cell(row=HDR_ROW + 1, column=col, value=hint)
        hint_cell.font      = Font(italic=True, size=9, color="94A3B8")
        hint_cell.alignment = left
        hint_cell.border    = border
        hint_cell.fill      = PatternFill("solid", fgColor="F8FAFC")

        ws.column_dimensions[get_column_letter(col)].width = width

    # Contoh data
    examples = [
        ["Senin",  "J.Int.1",  "1-3",  "TIF3221308", "Logika dan Himpunan",       "Dr. Budi Santoso, M.T.", "A,B",    ""],
        ["",       "J0403",    "1-3",  "TIF3221308", "Logika dan Himpunan",       "Siti Rahayu, M.Sc.",     "C,D",    ""],
        ["",       "LABRPL",   "7-9",  "TIF3232209", "Pemrograman Mobile",        "Dewi Kusumawati, M.Cs.", "A",      "https://classroom.google.com/c/abc"],
        ["Selasa", "J0407",    "1-3",  "SI2234567",  "Sistem Informasi Manajemen","Nur Aisyah, M.M.",       "A",      ""],
    ]

    for i, row_data in enumerate(examples):
        for col, val in enumerate(row_data, start=1):
            cell           = ws.cell(row=HDR_ROW + 2 + i, column=col, value=val)
            cell.alignment = left
            cell.border    = border
            cell.fill      = PatternFill("solid", fgColor="EFF6FF" if i % 2 == 0 else "FFFFFF")

    ws.row_dimensions[HDR_ROW].height = 22
    ws.freeze_panes = f"A{HDR_ROW + 2}"

    output = _io.BytesIO()
    wb.save(output)
    return output.getvalue()


# ─── MAIN ENTRY POINT ─────────────────────────────────────────

def parse_jadwal_file(file_bytes: bytes, filename: str) -> list[dict]:
    """
    Entry point utama — deteksi format dan parse.
    Return: list of parsed dicts
    """
    ext = filename.rsplit('.', 1)[-1].lower() if '.' in filename else ''

    if ext == 'pdf':
        return parse_pdf_jadwal(file_bytes)
    elif ext in ('xlsx', 'xls'):
        return parse_excel_jadwal(file_bytes)
    else:
        # Coba detect dari isi file
        if file_bytes[:4] == b'%PDF':
            return parse_pdf_jadwal(file_bytes)
        else:
            return parse_excel_jadwal(file_bytes)