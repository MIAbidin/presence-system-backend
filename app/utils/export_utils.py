# app/utils/export_utils.py
import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from typing import List
from app.models.presensi import Presensi


def generate_excel_rekap(
    presensi_list   : List[Presensi],
    nama_matakuliah : str,
    pertemuan_ke    : int,
    nama_dosen      : str = "",
    tanggal_sesi    : datetime | None = None,
    mode_kelas      : str = "",
) -> bytes:
    """
    Generate file Excel rekap presensi satu sesi.
    Return bytes — langsung bisa dikirim sebagai file response.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = f"Pertemuan {pertemuan_ke}"

    # ── Style ───────────────────────────────────────────────
    header_font    = Font(bold=True, color="FFFFFF", size=11)
    header_fill    = PatternFill("solid", fgColor="1E3A5F")
    center         = Alignment(horizontal="center", vertical="center")
    left           = Alignment(horizontal="left",   vertical="center")
    thin           = Side(style="thin", color="CBD5E1")
    border         = Border(left=thin, right=thin, top=thin, bottom=thin)

    STATUS_COLOR = {
        "hadir"    : "DCFCE7",
        "terlambat": "FEF3C7",
        "absen"    : "FEE2E2",
        "izin"     : "DBEAFE",
        "sakit"    : "F3E8FF",
    }

    # ── Info header (baris 1–4) ─────────────────────────────
    meta = [
        ("Matakuliah"  , nama_matakuliah),
        ("Pertemuan ke", str(pertemuan_ke)),
        ("Dosen"       , nama_dosen),
        ("Tanggal"     , tanggal_sesi.strftime("%d %B %Y") if tanggal_sesi else "-"),
        ("Mode"        , mode_kelas.upper() if mode_kelas else "-"),
    ]
    for i, (label, val) in enumerate(meta, start=1):
        ws.cell(row=i, column=1, value=label).font = Font(bold=True, size=10)
        ws.cell(row=i, column=2, value=val).font   = Font(size=10)

    # Judul besar
    ws.merge_cells("D1:H1")
    c = ws["D1"]
    c.value     = f"REKAP PRESENSI — {nama_matakuliah.upper()}"
    c.font      = Font(bold=True, size=14, color="1E3A5F")
    c.alignment = center

    ws.merge_cells("D2:H2")
    c2 = ws["D2"]
    c2.value     = f"Pertemuan {pertemuan_ke}  •  Mode: {mode_kelas.upper()}  •  {tanggal_sesi.strftime('%d %B %Y') if tanggal_sesi else ''}"
    c2.font      = Font(size=11, italic=True, color="64748B")
    c2.alignment = center

    # ── Header tabel (baris 7) ──────────────────────────────
    HDR_ROW = 7
    headers = ["No", "NIM", "Nama Mahasiswa", "Status", "Waktu Presensi", "Akurasi Wajah", "Mode", "Keterangan"]
    for col, h in enumerate(headers, start=1):
        cell           = ws.cell(row=HDR_ROW, column=col, value=h)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = center
        cell.border    = border

    # ── Data ────────────────────────────────────────────────
    for i, p in enumerate(presensi_list, start=1):
        row        = HDR_ROW + i
        status_str = p.status.value if p.status else "-"
        fill_color = STATUS_COLOR.get(status_str, "FFFFFF")
        row_fill   = PatternFill("solid", fgColor=fill_color)

        values = [
            i,
            p.mahasiswa.nim_nidn     if p.mahasiswa else "-",
            p.mahasiswa.nama_lengkap if p.mahasiswa else "-",
            status_str.upper(),
            p.waktu_presensi.strftime("%H:%M:%S") if p.waktu_presensi else "-",
            f"{p.akurasi_wajah:.1f}%" if p.akurasi_wajah else "-",
            (p.mode_kelas.value if p.mode_kelas else "-").upper(),
            p.catatan or "",
        ]

        aligns = [center, center, left, center, center, center, center, left]

        for col, (val, aln) in enumerate(zip(values, aligns), start=1):
            cell           = ws.cell(row=row, column=col, value=val)
            cell.border    = border
            cell.alignment = aln
            cell.fill      = row_fill

    # ── Ringkasan statistik ──────────────────────────────────
    summary_row = HDR_ROW + len(presensi_list) + 2
    counts = {
        "hadir"    : sum(1 for p in presensi_list if p.status and p.status.value == "hadir"),
        "terlambat": sum(1 for p in presensi_list if p.status and p.status.value == "terlambat"),
        "absen"    : sum(1 for p in presensi_list if p.status and p.status.value == "absen"),
        "izin"     : sum(1 for p in presensi_list if p.status and p.status.value == "izin"),
        "sakit"    : sum(1 for p in presensi_list if p.status and p.status.value == "sakit"),
    }
    total   = len(presensi_list)
    efektif = counts["hadir"] + counts["terlambat"]
    persen  = round(efektif / total * 100, 1) if total > 0 else 0.0

    summary_label_font = Font(bold=True, size=10)
    ws.cell(row=summary_row,   column=1, value="RINGKASAN").font = summary_label_font
    ws.cell(row=summary_row+1, column=1, value="Total Mahasiswa").font = Font(size=10)
    ws.cell(row=summary_row+1, column=2, value=total)
    for offset, (key, val) in enumerate(counts.items(), start=2):
        ws.cell(row=summary_row+offset, column=1, value=key.capitalize()).font = Font(size=10)
        ws.cell(row=summary_row+offset, column=2, value=val)
    ws.cell(row=summary_row+7, column=1, value="Persentase Hadir").font = Font(bold=True, size=10)
    ws.cell(row=summary_row+7, column=2, value=f"{persen}%").font       = Font(bold=True, size=10)

    # ── Lebar kolom ─────────────────────────────────────────
    col_widths = [5, 16, 30, 12, 18, 15, 10, 25]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Row heights
    ws.row_dimensions[HDR_ROW].height = 22

    # ── Freeze panes ────────────────────────────────────────
    ws.freeze_panes = f"A{HDR_ROW + 1}"

    # ── Simpan ke bytes ──────────────────────────────────────
    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()